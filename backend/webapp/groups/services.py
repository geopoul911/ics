from webapp.models import (
    GroupTransfer,
    History,
    TravelDay,
    Hotel,
    Service,
    Airline,
    DMC,
    FerryTicketAgency,
    CruisingCompany,
    Guide,
    Restaurant,
    SportEventSupplier,
    TeleferikCompany,
    Theater,
    TrainTicketAgency,
    Contact,
    CoachOperator,
    EntertainmentSupplier,
    EntertainmentProduct,
    Agent,
    Airport,
    Client,
)

from django.views.decorators.csrf import csrf_exempt
import openpyxl
from rest_framework.authtoken.models import Token
import os
from rest_framework import generics, status
from rest_framework.response import Response
from webapp.serializers import (
    GroupSerializer,
)
from bs4 import BeautifulSoup
from pathlib import Path
import datetime
from accounts.permissions import (
    can_create,
    can_update,
    can_delete,
)

BASE_DIR = Path(__file__).resolve().parent.parent.parent

SERVICE_TYPES = {
    'Accommodation': 'AC',
    'Air Ticket': 'AT',
    'Airport Porterage': 'AP',
    "Coach's Ferry Ticket": "CFT",
    'Cruise': 'CR',
    'Driver Accommodation': 'DA',
    'Ferry Ticket': 'FT',
    'Hotel Porterage': 'HP',
    'Local Guide': 'LG',
    'Restaurant': 'RST',
    'Shows & Entertainment': 'ES',
    'Sport Event': 'SE',
    'Sightseeing': 'SHT',
    'Teleferik': 'TE',
    'Theater': 'TH',
    'Tolls': 'TO',
    'Tour Leader': 'TL',
    "Tour Leader's Accommodation": 'TLA',
    "Tour Leader's Air Ticket": 'TLAT',
    'Train Ticket': 'TT',
    'Transfer': 'TR',
    'Other': 'OTH',
    'Permit': "PRM",
}

# Service types reversed
SERVICE_TYPES_REV = {
    "AC": 'Accommodation',
    "AT": 'Air Ticket',
    "AP": 'Airport Porterage',
    "CFT": "Coach's Ferry Ticket",
    "CR": 'Cruise',
    "DA": 'Driver Accommodation',
    "FT": 'Ferry Ticket',
    "HP": 'Hotel Porterage',
    "LG": 'Local Guide',
    "RST": 'Restaurant',
    'ES': 'Shows & Entertainment',
    "SE": 'Sport Event',
    'SHT': 'Sightseeing',
    "TE": 'Teleferik',
    "TH": 'Theater',
    "TO": 'Toll',
    "TL": 'Tour Leader',
    "TLA": "Tour Leader's Accommodation",
    "TLAT": "Tour Leader's Air Ticket",
    "TT": 'Train Ticket',
    "TR": 'Transfer',
    "OTH": 'Other',
    "PRM": 'Permit',
}

HOST_OPTIONS = {
    "Agent": "AG",
    "Airline": "AL",
    "Airport": "AP",
    "Client": "CL",
    "Cruising Company": "CC",
    "DMC": "DM",
    "Entertainment Supplier": "ES",
    "Ferry Ticket Agency": "FT",
    "Group Leader": "GL",
    "Guide": "GD",
    "Hotel": "HT",
    "Restaurant": "RT",
    "Sport Event Supplier": "SE",
    "Teleferik Company": "TC",
    "Theater": "TH",
    "Train Ticket Agency": "TT",
}


# Get user from token
def get_user(token):
    return Token.objects.get(key=token).user


def check_age(date_string):
    # Convert date string to datetime object
    date_object = datetime.datetime.strptime(date_string, '%Y-%m-%d')

    # Calculate age
    today = datetime.datetime.today()
    age = today.year - date_object.year - ((today.month, today.day) < (date_object.month, date_object.day))

    # Determine the age category
    if age <= 3:
        return "Infant"
    elif 3 < age <= 15:
        return "Child"
    else:
        return "Adult"


# Adjust column widths
column_widths = {
    'A': 100 // 7,
    'B': 100 // 7,
    'C': 100 // 7,
    'D': 100 // 7,
    'E': 100 // 7,
    'F': 200 // 7,
    'G': 100 // 7,
    'H': 200 // 7
}


def create_cabin_list(service):
    td = TravelDay.objects.get(id=service.travelday_id)
    group = GroupTransfer.objects.get(id=td.group_transfer_id)

    folder_path = os.path.join(BASE_DIR, 'files', 'cabin_lists', group.refcode)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    wb = openpyxl.Workbook()
    ws = wb.active

    img = openpyxl.drawing.image.Image(str(BASE_DIR) + '/images/logos/stena_line_logo.png')
    img.anchor = 'F1'
    ws.add_image(img)

    data = [
        ['', '', ''],
        ['', '', ''],
        ['Please enter the required information below for every person travelling.', '', ''],
        ['Gender, Category, Nationality, Date of Birth plus Special Needs (if applicable) must be chosen by using the dropdown arrows.', '', ''],
        ['Please return the completed list to group.travel@stenaline.com at least 2 weeks prior to departure date.', '', ''],
        ['', '', ''],
        ['If your group is travelling by vehicle, please remember to enter the driver’s name.  If you do not know the driver’s', '', ''],
        ["name, please enter 'driver' in both the forename and surname boxes (example below).", '', ''],
        ['', '', ''],
        ['Booking Ref: ', service.booking_ref],
        ['Route: ', service.route],
        ['Date: ', service.date],
        ['', '', ''],
        ['#', 'First name', 'Surname', 'Gender', 'Category', 'Nationality', 'Special needs', 'Date of Birth (dd/mm/yyyy)']
    ]

    # Adding data to the worksheet
    for row in data:
        ws.append(row)

    # Apply styling to header row
    for cell in ws[14]:
        cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")  # Grey background
        cell.font = openpyxl.styles.Font(color="0000FF", bold=True)  # Blue bold font

    # Parsing HTML table
    soup = BeautifulSoup(group.roomtext, 'html.parser')
    table = soup.find('table')
    rows = table.find_all('tr')[1:]

    rowspan_tracker = {}

    for row_index, row in enumerate(rows, start=1):
        dob = ''
        cols = row.find_all('td')
        col_index = 0
        excel_row = []
        for col in cols:
            while (row_index, col_index) in rowspan_tracker:
                excel_row.append(rowspan_tracker[(row_index, col_index)])
                col_index += 1

            rowspan = int(col.get('rowspan', 1))
            cell_value = col.text.strip()

            for i in range(rowspan):
                rowspan_tracker[(row_index + i, col_index)] = cell_value

            """
            GROUP'S TABLE INDEXES:
            1 = Room Type
            2 = ID
            3 = Full Name
            4 = Gender
            5 = Date Of Birth
            6 = Passport
            """

            if col_index == 2:
                if "/" in cell_value:  # Check if the cell contains both first and last names
                    first_name, last_name = cell_value.split("/")  # Split into first and last name
                    excel_row.append(first_name.strip())  # Append first name
                    excel_row.append(last_name.strip())  # Append last name
                else:
                    excel_row.append(cell_value)

            elif col_index == 3:
                if cell_value == 'M':
                    cell_value = 'Male'
                elif cell_value == 'F':
                    cell_value = 'Female'

            elif col_index == 4 and cell_value != '':
                dob = datetime.datetime.strptime(cell_value, '%Y-%m-%d').strftime('%d/%m/%Y')
                cell_value = check_age(cell_value)

            elif col_index == 5:
                if group.agent.nationality.name == 'China':
                    cell_value = 'China (Peoples Republic Of)'
                else:
                    cell_value = group.agent.nationality.name

            if col_index != 2:
                excel_row.append(cell_value)
            col_index += 1

        excel_row.append('')
        excel_row.append(dob)
        ws.append(excel_row[1:])

        # Define border style
        border_style = openpyxl.styles.Border(
            left=openpyxl.styles.Side(border_style='thin'),
            right=openpyxl.styles.Side(border_style='thin'),
            top=openpyxl.styles.Side(border_style='thin'),
            bottom=openpyxl.styles.Side(border_style='thin')
        )

        # Apply borders to the row
        current_row = ws[row_index + 13]  # Adjusting for 0-based indexing and header row
        for cell in current_row:
            cell.border = border_style

    bold_font = openpyxl.styles.Font(bold=True)
    for cell in ['A10', 'A11', 'A12']:
        ws[cell].font = bold_font
        ws[cell].border = border_style

    ws.column_dimensions["A"].width = column_widths["A"]
    ws.column_dimensions["B"].width = column_widths["B"]
    ws.column_dimensions["C"].width = column_widths["C"]
    ws.column_dimensions["D"].width = column_widths["D"]
    ws.column_dimensions["E"].width = column_widths["E"]
    ws.column_dimensions["F"].width = column_widths["F"]
    ws.column_dimensions["G"].width = column_widths["G"]
    ws.column_dimensions["H"].width = column_widths["H"]

    ws.row_dimensions[14].height = 50
    for col in range(1, ws.max_column + 1):
        ws.cell(row=14, column=col).alignment = openpyxl.styles.Alignment(vertical='center')

    # Iterate over cells in the last row starting from the fourth cell
    for col in range(4, ws.max_column + 1):
        # Clear the contents of the cell
        ws.cell(row=ws.max_row, column=col).value = None

    # Iterate over cells in the last row starting from the fourth cell
    for col in range(1, ws.max_column + 1):
        # Apply border style to the cell
        ws.cell(row=ws.max_row, column=col).border = border_style

    file_path = os.path.join(BASE_DIR, 'files', 'cabin_lists', group.refcode, f'CabinList_{group.refcode}_{service.booking_ref}.xlsx')
    wb.save(file_path)

    return


class DeleteService(generics.ListCreateAPIView):
    """
    url : delete_service/(?P<refcode>.*)$
    Descr: Deletes a service
    """

    # Cross-Site Request Forgery
    @csrf_exempt
    def post(self, request, refcode):
        token_str = request.headers['User-Token']
        user = get_user(token_str)
        context = {"errormsg": ''}

        # Permission
        if not can_delete(user.id, 'GT'):
            context = {"errormsg": "You do not have permission to delete a Group's service."}
            return Response(status=401, data=context)

        # Get service ID
        service_id = request.data['service_id']
        service = Service.objects.get(id=service_id)

        # Get Group
        group = GroupTransfer.objects.get(refcode=refcode)

        # Get Travelday
        travelday = TravelDay.objects.get(id=service.travelday_id)

        # Get Service
        try:
            service_to_del = Service.objects.get(id=service_id)
            History.objects.create(
                user=user,
                model_name='SRV',
                action='DEL',
                description=f"User : {user.username} deleted group's {group.refcode} \
                    service: {SERVICE_TYPES_REV[service_to_del.service_type]} with date : \
                    {travelday.date} and id: {service_id}"
            )
            service_to_del.delete()
        except Exception as a:
            context['errormsg'] = a
            return Response(data=context, status=400)

        if group == 'N/A':
            return Response(data=context, status=200)

        context['model'] = GroupSerializer(group).data
        return Response(data=context, status=200)


class AddNewService(generics.UpdateAPIView):
    """
    URL: create_travelday_service/$
    Descr: Creates travelday service.
    """

    @csrf_exempt
    def post(self, request):
        try:
            token_str = request.headers['User-Token']
            user = get_user(token_str)
            context = {"errormsg": ''}

            if not can_create(user.id, 'SRV'):
                context = {"errormsg": "You do not have permission to create a Service."}
                return Response(status=401, data=context)

            service_type = request.data.get('service_type')
            description = request.data.get('description')
            price = request.data.get('price')
            currency = request.data.get('currency')
            time = request.data.get('time')

            route = request.data.get('route')
            booking_ref = request.data.get('booking_ref')

            try:
                host_name = request.data.get('host_name')
                host_id = None
                if request.data.get('host_type') == "Agent":
                    host_id = Agent.objects.get(name=host_name).id
                elif request.data.get('host_type') == "Airline":
                    host_id = Airline.objects.get(name=host_name).id
                elif request.data.get('host_type') == "Airport":
                    host_id = Airport.objects.get(name=host_name).name
                elif request.data.get('host_type') == "Client":
                    host_id = Client.objects.get(name=host_name).id
                elif request.data.get('host_type') == "Cruising Company":
                    host_id = CruisingCompany.objects.get(name=host_name).id
                elif request.data.get('host_type') == "DMC":
                    host_id = DMC.objects.get(name=host_name).id
                elif request.data.get('host_type') == "Entertainment Supplier":
                    host_id = EntertainmentSupplier.objects.get(name=host_name).id
                elif request.data.get('host_type') == "Ferry Ticket Agency":
                    host_id = FerryTicketAgency.objects.get(name=host_name).id
                elif request.data.get('host_type') == "Group Leader":
                    host_id = Contact.objects.get(name=host_name, type='L').id
                elif request.data.get('host_type') == "Guide":
                    host_id = Guide.objects.get(name=host_name).id
                elif request.data.get('host_type') == "Hotel":
                    host_id = Hotel.objects.get(name=host_name).id
                elif request.data.get('host_type') == "Restaurant":
                    host_id = Restaurant.objects.get(name=host_name).id
                elif request.data.get('host_type') == "Sport Event Supplier":
                    host_id = SportEventSupplier.objects.get(name=host_name).id
                elif request.data.get('host_type') == "Teleferik Company":
                    host_id = TeleferikCompany.objects.get(name=host_name).id
                elif request.data.get('host_type') == "Theater":
                    host_id = Theater.objects.get(name=host_name).id
                elif request.data.get('host_type') == "Train Ticket Agency":
                    host_id = TrainTicketAgency.objects.get(name=host_name).id
            except Exception as a:
                print(a)
                host_id = None

            try:
                host_type = HOST_OPTIONS[request.data.get('host_type')]
            except Exception as a:
                print(a)
                host_type = None

            try:
                ent_product = EntertainmentProduct.objects.get(name=request.data.get('entertainment_product'))
            except Exception:
                ent_product = None

            # Handle dates
            if request.data.get('way') == 'data_management':
                dates = {TravelDay.objects.get(id=request.data['td_id']).date: True}
            else:
                dates = request.data.get('dates')

            # Initialize variables
            hotel = airline = dmc = ferry_ticket_agency = cruising_company = None
            guide = sport_event_supplier = teleferik_company = theater = None
            train_ticket_agency = tour_leader = coach_operator = None
            restaurant = entertainment_supplier = None

            if service_type == 'Accommodation' or service_type == "Tour Leader's Accommodation":
                hotel = Hotel.objects.get(name=request.data.get('hotel'))
            elif service_type == 'Air Ticket' or service_type == "Tour Leader's Air Ticket":
                airline = Airline.objects.get(name=request.data.get('airline'))
            elif service_type == 'Airport Porterage':
                dmc = DMC.objects.get(name=request.data.get('dmc'))
            elif service_type == "Coach's Ferry Ticket" or service_type == 'Ferry Ticket':
                ferry_ticket_agency = FerryTicketAgency.objects.get(name=request.data.get('ferry_ticket_agency'))
            elif service_type == 'Cruise':
                cruising_company = CruisingCompany.objects.get(name=request.data.get('cruising_company'))
            elif service_type == 'Local Guide':
                guide = Guide.objects.get(name=request.data.get('guide'))
            elif service_type == 'Restaurant':
                restaurant = Restaurant.objects.get(name=request.data.get('restaurant'))
            elif service_type == 'Sport Event':
                sport_event_supplier = SportEventSupplier.objects.get(name=request.data.get('sport_event_supplier'))
            elif service_type == 'Teleferik':
                teleferik_company = TeleferikCompany.objects.get(name=request.data.get('teleferik_company'))
            elif service_type == 'Theater':
                theater = Theater.objects.get(name=request.data.get('theater'))
            elif service_type == 'Train Ticket':
                train_ticket_agency = TrainTicketAgency.objects.get(name=request.data.get('train_ticket_agency'))
            elif service_type == 'Tour Leader':
                tour_leader = Contact.objects.get(name=request.data.get('tour_leader'))
            elif service_type == 'Transfer':
                coach_operator = CoachOperator.objects.get(name=request.data.get('coach_operator'))
            elif service_type == 'Shows & Entertainment':
                entertainment_supplier = EntertainmentSupplier.objects.get(name=request.data.get('entertainment_supplier'))

            group = GroupTransfer.objects.get(refcode=request.data.get('refcode'))

            for date, value in dict(dates).items():
                if value:
                    try:
                        travelday_id = TravelDay.objects.get(date=date, group_transfer_id=group.id).id
                    except TravelDay.MultipleObjectsReturned:
                        travelday_id = TravelDay.objects.filter(date=date, group_transfer_id=group.id)[0].id

                    new_service = Service.objects.create(
                        travelday_id=travelday_id,
                        service_type=SERVICE_TYPES.get(service_type),
                        description=description,
                        price=price,
                        currency=currency,
                        date=date,
                        hotel=hotel,
                        airline=airline,
                        dmc=dmc,
                        ferry_ticket_agency=ferry_ticket_agency,
                        cruising_company=cruising_company,
                        guide=guide,
                        restaurant=restaurant,
                        sport_event_supplier=sport_event_supplier,
                        teleferik_company=teleferik_company,
                        theater=theater,
                        train_ticket_agency=train_ticket_agency,
                        tour_leader=tour_leader,
                        coach_operator=coach_operator,
                        start_time=time,
                        route=route,
                        booking_ref=booking_ref,
                        entertainment_supplier=entertainment_supplier,
                        entertainment_product=ent_product,
                        host_id=host_id,
                        host_type=host_type,
                    )

                    History.objects.create(
                        user=user,
                        model_name='SRV',
                        action='CRE',
                        description=f"User : {user.username} created a service: {SERVICE_TYPES_REV.get(new_service.service_type)} \
                            with id: {new_service.id} for group: {group.refcode}"
                    )

                    new_service.save()

            context['model'] = GroupSerializer(group).data
            return Response(data=context, status=200)

        except Exception as e:
            context['errormsg'] = str(e)
            return Response(data=context, status=400)


class ChangeDate(generics.UpdateAPIView):
    """
    URL: change_date/
    Descr: Updates service's date
    """

    # Cross-Site Request Forgery
    @csrf_exempt
    def post(self, request):
        context = {"errormsg": ''}
        token_str = request.headers['User-Token']
        user = get_user(token_str)

        # Permission
        if not can_update(user.id, 'SRV'):
            context = {"errormsg": "You do not have permission to update Service's date."}
            return Response(status=401, data=context)

        # Get service
        service = Service.objects.get(id=request.data['service_id'])

        # Get previous td for logging
        previous_td = TravelDay.objects.get(id=service.travelday_id)

        # Get group
        group = GroupTransfer.objects.get(id=previous_td.group_transfer_id)

        # Get new date
        new_date = request.data['selected_date']
        selected_day = TravelDay.objects.get(group_transfer_id=group.id, date=new_date)

        try:
            service.travelday_id = selected_day.id
            service.date = selected_day.date
            service.save()

            History.objects.create(
                user=user,
                model_name='SRV',
                action='UPD',
                description=f"User {user} updated group's: {group.refcode} service's (id: {service.id}) date \
                    from {previous_td.date} to {selected_day.date}"
            )

            # Serialize data
            context['model'] = GroupSerializer(group).data
            return Response(data=context, status=status.HTTP_200_OK)
        except Exception as exc:
            context['errormsg'] = exc
            return Response(data=context, status=400)


class ChangePrice(generics.UpdateAPIView):
    """
    URL: change_price/
    Descr: Updates service's price
    """

    # Cross-Site Request Forgery
    @csrf_exempt
    def post(self, request):
        context = {"errormsg": ''}
        token_str = request.headers['User-Token']
        user = get_user(token_str)

        # Permission
        if not can_update(user.id, 'SRV'):
            context = {"errormsg": "You do not have permission to update Service's price."}
            return Response(status=401, data=context)

        # Get service
        service = Service.objects.get(id=request.data['service_id'])
        price = request.data['price']
        prev_price = service.price

        currency = request.data['currency']
        prev_currency = service.currency

        # Get previous td for logging
        previous_td = TravelDay.objects.get(id=service.travelday_id)

        # Get group
        group = GroupTransfer.objects.get(id=previous_td.group_transfer_id)

        try:
            service.price = price
            service.currency = currency
            service.save()

            History.objects.create(
                user=user,
                model_name='SRV',
                action='UPD',
                description=f"User {user} updated group's: {group.refcode} service's (id: {service.id}) date \
                    from {prev_currency} {prev_price} to {currency} {price}"
            )

            # Serialize data
            context['model'] = GroupSerializer(group).data
            return Response(data=context, status=status.HTTP_200_OK)
        except Exception as exc:
            context['errormsg'] = exc
            return Response(data=context, status=400)


class ChangeDescription(generics.UpdateAPIView):
    """
    URL: change_description/
    Descr: Updates service's description
    """

    # Cross-Site Request Forgery
    @csrf_exempt
    def post(self, request):
        context = {"errormsg": ''}
        token_str = request.headers['User-Token']
        user = get_user(token_str)

        # Permission
        if not can_update(user.id, 'SRV'):
            context = {"errormsg": "You do not have permission to update Service's description."}
            return Response(status=401, data=context)

        # Get description
        description = request.data['description']

        # Get service
        service = Service.objects.get(id=request.data['service_id'])

        # Get previous description for logging
        prev_description = service.description

        group = GroupTransfer.objects.get(id=TravelDay.objects.get(id=service.travelday_id).group_transfer_id)

        try:
            service.description = description
            service.save()
            History.objects.create(
                user=user,
                model_name='SRV',
                action='UPD',
                description=f'User {user} updated service\'s description from {prev_description} to {description}'
            )

            # Serialize data
            context['model'] = GroupSerializer(group).data
            return Response(data=context, status=status.HTTP_200_OK)
        except Exception as exc:
            context['errormsg'] = exc
            return Response(data=context, status=400)


class DeleteAllServices(generics.UpdateAPIView):
    """
    URL : delete_all_services/(?P<refcode>.*)$
    Descr: Deletes selected group's services based on refcode
    """

    # Cross-Site Request Forgery
    @csrf_exempt
    def post(self, request, refcode):
        context = {"errormsg": ''}
        token_str = request.headers['User-Token']
        user = get_user(token_str)

        # Permissions
        if not can_delete(user.id, 'GT'):
            context = {"errormsg": "You do not have permission to delete a Group's services."}
            return Response(status=401, data=context)

        # Get group
        group = GroupTransfer.objects.get(refcode=refcode)
        try:
            # Get traveldays
            traveldays_to_delete = TravelDay.objects.filter(group_transfer_id=group.id)
            td_ids = [td.id for td in traveldays_to_delete]

            # Get services to delete
            services_to_delete = Service.objects.filter(travelday_id__in=td_ids)

            # Log all deletes
            for service in services_to_delete:
                History.objects.create(
                    user=user,
                    model_name='GT',
                    action='DEL',
                    description=f"User : {user.username} deleted service with id: {service.id} \
                        on date: {TravelDay.objects.get(id=service.travelday_id).date} of group: {group.refcode}"
                )
                service.delete()
            context['model'] = GroupSerializer(group).data
        except Exception as a:
            context['errormsg'] = a
            return Response(data=context, status=400)
        return Response(data=context, status=200)
